from pathlib import Path
from typing import *
from fastapi import FastAPI, HTTPException, UploadFile, File
from pydantic import BaseModel
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment

import database
import ai_core.api as ai_api

import database.accounts_db as accounts_db
import database.dataset_db as dataset_db
import database.schedule_db as schedule_db

app = FastAPI()

MAX_MATCHES = 22 # takes into account: 13 min meetings, with 2 mins in between, with two 7 minutes breaks, over 2 days from 3pm-6pm

# ---------- Matchmaking: normalize raw AI scores to user-friendly percentages ----------
# Raw cosine similarity is typically 0.2–0.7. We map top score -> 100% and filter out
# results that are "too far" from the top so the list looks consistent.
DEFAULT_MATCH_THRESHOLD_PERCENT = 50.0   # minimum normalized % to include
MIN_RELATIVE_TO_TOP = 0.7                # drop if raw_score < top_raw * 0.7


def _normalize_and_filter_matches(raw_results: list, match_threshold_percent: Optional[float]) -> list:
    """
    raw_results: list of {"ngo_id", "ngo", "score"} sorted by score desc.
    Returns list with ai_match_score_percent (0–100), and excludes matches too far from top.
    """
    if not raw_results:
        return []
    threshold = match_threshold_percent if match_threshold_percent is not None else DEFAULT_MATCH_THRESHOLD_PERCENT
    top_raw = raw_results[0]["score"]
    if top_raw <= 0:
        top_raw = 1e-9
    out = []
    for r in raw_results:
        raw = r["score"]
        if raw < top_raw * MIN_RELATIVE_TO_TOP:
            continue
        pct = round((raw / top_raw) * 100.0, 1)
        if pct < threshold:
            continue
        ngo_dict = dict(r["ngo"]) if hasattr(r["ngo"], "keys") else r["ngo"]
        out.append({
            "ngo_id": r["ngo_id"],
            "ngo_name": ngo_dict.get("name") or "",
            "ngo_strategy": (ngo_dict.get("strategy") or ngo_dict.get("focus") or ""),
            "raw_score": raw,
            "ai_match_score_percent": pct,
            "possible_reasons": "",  # optional: could derive from strategy overlap later
        })
    return out

@app.on_event("startup")
async def on_startup() -> None:
    # Ensure the SQLite databases exist before handling any requests
    database.initialize_all_databases()
    # No automatic mock data – use import scripts or API import-excel for data

# 1. Define the blueprint for the incoming data
class LoginData(BaseModel):
    email: str
    password: str

# 2. Create the POST endpoint to receive the data
@app.post("/api/login")
async def process_login(data: LoginData):
    # Endpoint to handle login requests from the frontend
    is_valid = accounts_db.verify_credentials(data.email, data.password)

    print(f"Login attempt for email: {data.email} - {'SUCCESS' if is_valid else 'FAILURE'}")
    
    if is_valid:
        return {"status": "success", "message": "Welcome to HIGGS AI matchmaking platform!"}
    else:
        # FastAPI's standard way to handle errors
        raise HTTPException(status_code=401, detail="Invalid email or password")
    

@app.post("/api/ensure-embeddings")
async def ensure_embeddings():
    """
    Precompute embeddings for all donors and NGOs that don't have one.
    Call this after importing data so the first recommendations request is fast.
    """
    ai_api.ensure_embeddings()
    return {"status": "ok", "message": "Embeddings updated for all donors and NGOs."}


@app.get("/api/donors/{donor_id}/recommendations")
async def get_donor_recommendations(
    donor_id: int,
    top_k: int = 10,
    save_matches: bool = False,
):
    """
    Return top NGO recommendations for a donor by semantic similarity.
    top_k: max number of results (default 10).
    save_matches: if true, store scores in donor_ngo_matches table.
    """
    if database.dataset_db.get_donor(donor_id) is None:
        raise HTTPException(status_code=404, detail="Donor not found")
    results = ai_api.get_recommendations_for_donor(
        donor_id=donor_id,
        top_k=top_k,
        save_matches=save_matches,
    )
    return {"donor_id": donor_id, "recommendations": results}

class RegisterData(BaseModel):
    email: str
    password: str

@app.post("/api/register")
async def register_account(data: RegisterData):
    try:
        account_id = accounts_db.create_account(
            data.email,
            data.password,
            approved=True, # for testing
            verified=True, # for testing
            is_admin=False,
        )
        return {"status": "success", "message": f"Account created with ID {account_id}"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

def _row_to_jsonable(row) -> dict:
    """Convert a DB row (donor or NGO) to a dict, omitting binary 'embedding' so JSON encoding works."""
    d = dict(row)
    d.pop("embedding", None)
    return d


@app.get("/api/donors")
async def get_all_donors():
    donors = dataset_db.list_donors()
    return [_row_to_jsonable(d) for d in donors]

@app.get("/api/ngos")
async def get_all_ngos():
    ngos = dataset_db.list_ngos()
    return [_row_to_jsonable(n) for n in ngos]


@app.get("/api/getmatches/{donor_id}")
async def get_donor_matches(donor_id: int):
    matches = dataset_db.list_matches_for_donor(donor_id=donor_id)
    return [_row_to_jsonable(m) for m in matches]

@app.get("/api/getmatches/all")
async def get_all_matchs():
    matches = dataset_db.list_all_matches()
    return [_row_to_jsonable(m) for m in matches]

# ---------- Add / delete NGOs (single donor matchmaking; NGOs managed separately) ----------
# Donors and NGOs live in the same database (dataset.db) but in different tables.
# Users can add/delete NGOs without affecting donors.

class AddNGOBody(BaseModel):
    name: str
    strategy: Optional[str] = None
    focus: Optional[str] = None
    legal_form: Optional[str] = None


@app.post("/api/ngos")
async def add_ngo(body: AddNGOBody):
    """Add a single NGO to the database. Used by 'Add NGOs' (e.g. upload list)."""
    if not (body.name or "").strip():
        raise HTTPException(status_code=400, detail="NGO name is required")
    ngo_id = dataset_db.insert_ngo(
        name=body.name.strip(),
        strategy=body.strategy,
        focus=body.focus or body.strategy,
        legal_form=body.legal_form,
    )
    return {"status": "ok", "id": ngo_id, "message": f"NGO '{body.name.strip()}' added."}


class DeleteNGOsBody(BaseModel):
    ids: list[int]


@app.delete("/api/ngos")
async def delete_ngos(body: DeleteNGOsBody):
    """Delete a list of NGOs by ID. Used by 'Delete NGOs'."""
    if not body.ids:
        raise HTTPException(status_code=400, detail="ids must be a non-empty list")
    deleted = dataset_db.delete_ngos(body.ids)
    deleted_meetings = schedule_db.delete_many_ngo_meetings(body.ids)
    return {"status": "ok", "deleted": deleted, "deleted_meetings": deleted_meetings}


# ---------- Single-donor matchmaking (donor entered by user, not from DB) ----------

class MatchmakingGenerateBody(BaseModel):
    donor_name: str
    donor_strategy: str
    match_threshold_percent: Optional[float] = None  # e.g. 70 = only show matches >= 70%


@app.post("/api/matchmaking/generate")
async def matchmaking_generate(body: MatchmakingGenerateBody):
    """
    Generate AI matches for a single donor (name + strategy entered by user).
    Returns NGOs ranked by similarity, with normalized scores (0–100%) and filtered
    so that only matches close enough to the top score are included.
    """
    if not (body.donor_name or "").strip():
        raise HTTPException(status_code=400, detail="Donor name is required")
    if not (body.donor_strategy or "").strip():
        raise HTTPException(status_code=400, detail="Donor strategy is required")

    raw = ai_api.get_recommendations_for_donor_profile(
        donor_name=body.donor_name.strip(),
        donor_strategy=body.donor_strategy.strip(),
        top_k=100,
    )
    matches = _normalize_and_filter_matches(raw, body.match_threshold_percent)

    return {
        "donor_name": body.donor_name.strip(),
        "matches": matches,
        "total_matched": len(matches),
    }


class MatchmakingExportBody(BaseModel):
    donor_name: str
    donor_strategy: str
    match_threshold_percent: Optional[float] = None


# CSV export for matchmaking: saved here and overwritten on each export (desktop app reads this path)
EXPORT_DIR = Path(__file__).resolve().parents[1] / "data" / "export_file"
EXPORT_CSV_FILENAME = "matchmaking_results.csv"


@app.post("/api/matchmaking/export")
async def matchmaking_export(body: MatchmakingExportBody):
    """
    Run matchmaking, write results to data/export_file/matchmaking_results.csv (overwrites each time),
    and return the path for the desktop app to open. Same donor/strategy/threshold as generate.
    """
    if not (body.donor_name or "").strip():
        raise HTTPException(status_code=400, detail="Donor name is required")
    if not (body.donor_strategy or "").strip():
        raise HTTPException(status_code=400, detail="Donor strategy is required")

    raw = ai_api.get_recommendations_for_donor_profile(
        donor_name=body.donor_name.strip(),
        donor_strategy=body.donor_strategy.strip(),
        top_k=100,
    )
    matches = _normalize_and_filter_matches(raw, body.match_threshold_percent)
    # Only write top 5 (or fewer if threshold left less) so the file is short and clear
    export_matches = matches[:5]
    donor_name = body.donor_name.strip()

    import csv
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    export_path = EXPORT_DIR / EXPORT_CSV_FILENAME
    with export_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Donor", donor_name])
        writer.writerow(["#", "NGO name", "NGO strategy", "AI Match Score (%)", "Possible reasons"])
        for i, m in enumerate(export_matches, 1):
            writer.writerow([
                i,
                m["ngo_name"],
                m["ngo_strategy"],
                m["ai_match_score_percent"],
                m["possible_reasons"] or "",
            ])

    # Return path relative to project root so the desktop app can open the file
    relative_path = str(Path("data") / "export_file" / EXPORT_CSV_FILENAME)
    return {
        "saved_to": relative_path,
        "path_absolute": str(export_path),
        "filename": EXPORT_CSV_FILENAME,
        "donor_name": donor_name,
        "total_matched": len(matches),
        "exported_count": len(export_matches),
    }

@app.post("/api/import-excel")
async def import_excel(file: UploadFile = File(...)):
    """
    Upload an Excel file from the desktop app and import donors/NGOs into the DB.
    The file is stored under data/imports/ and then processed.
    """
    from pathlib import Path
    import shutil

    project_root = Path(__file__).resolve().parents[1]
    imports_dir = project_root / "data" / "imports"
    imports_dir.mkdir(parents=True, exist_ok=True)

    save_path = imports_dir / file.filename
    with save_path.open("wb") as f:
        shutil.copyfileobj(file.file, f)

    summary = database.excel_import.import_dataset_from_excel(save_path)
    return {"file": file.filename, "summary": summary}



@app.post("/api/download-matches-workbook")
async def export_matches_wk():
    matches = dataset_db.list_all_matches_details()
    matches_rows = [_row_to_jsonable(m) for m in matches]

    filename = "matchmaking_saved_results_" + datetime.now().strftime("%Y-%m-%d_%H-%M-%S")+ ".xlsx"
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    export_path = EXPORT_DIR / filename

    create_file(export_path, matches_rows)


""""
API for scheduling.
"""

class MeetingDate(BaseModel):
    donor_id: int
    ngo_id: int
    timestamp: datetime

@app.post("/api/schedule/add")
async def add_meeting(body: MeetingDate):
    try:
        new_id = schedule_db.insert_meeting(
                    body.donor_id,
                    body.ngo_id,
                    body.timestamp
                )
        return {"status": "success", "message": f"Meeting created with ID {new_id}"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    
@app.get("/api/schedule/donor/{donor_id}/meetings")
async def get_donor_meetings(donor_id: int):
    meetings = schedule_db.get_donor_meetings(donor_id=donor_id)
    return [_row_to_jsonable(m) for m in meetings]

@app.get("/api/schedule/ngo/{ngo_id}/meetings")
async def get_ngo_meetings(ngo_id: int):
    meetings = schedule_db.get_ngo_meetings(ngo_id=ngo_id)
    return [_row_to_jsonable(m) for m in meetings]

@app.get("/api/schedule/meeting/{donor_id}/{ngo_id}")
async def get_meeting(donor_id: int, ngo_id: int):
    meeting = schedule_db.get_donor_ngo_meeting(donor_id=donor_id, ngo_id=ngo_id)
    return meeting

class MeetingList(BaseModel):
    meetings: List[Tuple[int, int, str, datetime]] # (donor id, ngo id, meeting time)

@app.post("/api/schedule/add-many-meetings")
async def add_many_meetings(body: MeetingList):
    print("adding meetings api")
    try:
        schedule_db.batch_insert_meetings(body.meetings)
        return {"status": "success", "message": "Meetings created"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    
@app.get("/api/schedule/get-all-meetings")
async def get_all_meetings():
    meetings = schedule_db.get_all_meetings()
    return [_row_to_jsonable(m) for m in meetings]

# Add and delete donors

class AddDonorBody(BaseModel):
    name: str
    strategy: Optional[str] = None
    legal_form: Optional[str] = None

@app.post("/api/donors")
async def add_donor(body: AddDonorBody):
    if not (body.name or "").strip():
        raise HTTPException(status_code=400, detail="Donor name is required")
    donor_id = dataset_db.insert_donor(
        name=body.name.strip(),
        strategy=body.strategy,
        legal_form=body.legal_form,
    )
    return {"status": "ok", "id": donor_id, "message": f"Donor '{body.name.strip()}' added."}

class DeleteDonorsBody(BaseModel):
    ids: list[int]

@app.delete("/api/donors")
async def delete_donors(body: DeleteDonorsBody):
    if not body.ids:
        raise HTTPException(status_code=400, detail="ids must be a non-empty list")
    deleted = dataset_db.delete_donors(body.ids)
    deleted_meetings = schedule_db.delete_many_donor_meetings(body.ids)
    return {"status": "ok", "deleted": deleted, "deleted_meetings": deleted_meetings}

def create_file(name: str, data: list):
    workbook = Workbook()
    sheet = workbook.active

    # column names
    sheet.append([
        "#",
        "Donor",
        "Legal Form",
        "Strategy",
        "Organizations they picked",
        "Organizations Strategy",
        "Similarity"
    ])

    sheet.freeze_panes = "A2"

    column_widths = {
        1: 5,   # # (ID)
        2: 20,  # Donor
        3: 15,  # Legal Form
        4: 60,  # Strategy
        5: 20,  # Organizations
        6: 60,   # Organizations Strategy 
        7: 10    # Similarity
    }

    for col_num, width in column_widths.items():
        column_letter = sheet.cell(row=1, column=col_num).column_letter
        sheet.column_dimensions[column_letter].width = width

    """
    Data is rows of dicts. Need to parse into list of lists first.
    As the data is sorted by donor id ascending, this groups the same donors together. 
    We will check if the processing for each donor is done first, and then merge rows

    """
    if len(data) > 0:
        next_row = sheet.max_row + 1 # the next row 
        rows = [list(row.values()) for row in data]
        prev_row_id = rows[0][0]
        prev_row_num = next_row # keeps track of where to merge from
        for r in rows:
            sheet.append(r)
            if r[0] != prev_row_id and sheet.max_row > 3:
                print("merge")
                merge_to = sheet.max_row - 1
                merge_cells(prev_row_num, merge_to, sheet)

                prev_row_num = sheet.max_row

            prev_row_id = r[0]
        # last group
        if sheet.max_row > prev_row_num:
            merge_to = sheet.max_row
            merge_cells(prev_row_num, merge_to, sheet)

    wrap_cells(sheet)
    workbook.save(filename=name)


def merge_cells(prev_row_num, merge_to, sheet):
    for c in range(1,5):
        sheet.merge_cells(start_row=prev_row_num, start_column=c, end_row=merge_to, end_column=c)

def wrap_cells(sheet):
    style = Alignment(wrap_text=True, vertical="top", horizontal="left")

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = style