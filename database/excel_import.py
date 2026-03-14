"""
Import donors and NGOs from the HIGGS Excel format.

Structure (e.g. 2023 data):
- Row 1: year (e.g. "2023") – skip as header.
- Row 2: #, Donor, Legal form, Strategy, Organizations they picked, Organization strategy
- Donor and Legal form are MERGED: only first row of each block has value; rest are NaN.
- Strategy = donor's strategy (what they want to fund); one per row in block.
- Organizations they picked = one NGO name per row.
- Organization strategy = that NGO's focus/strategies (for the NGO, not the donor).

Result: exactly 17 donors (unique names), and one row per unique NGO with its focus.
"""

from pathlib import Path
from typing import Optional, Dict, Any
from io import BytesIO

import pandas as pd

from . import dataset_db


def import_dataset_from_excel(excel_path: str | Path) -> Dict[str, int]:
    """
    Read the HIGGS Excel and insert:
    - One row per unique Donor (name, legal form, aggregated strategies) → 17 donors for 2023.
    - One row per unique NGO (Organizations they picked + Organization strategy as focus).
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        return {"donors_inserted": 0, "ngos_inserted": 0}

    dataset_db.initialize_schema()

    try:
        df = _read_excel_with_merged_cells_filled(excel_path)
    except Exception:
        df = None
    if df is None or df.empty:
        df = _read_excel_fallback(excel_path)
    if df is None or df.empty:
        return {"donors_inserted": 0, "ngos_inserted": 0}

    df = _normalize_columns(df)
    col_map = _resolve_higgs_columns(df)
    # If openpyxl gave wrong shape, try pandas-only fallback (ffill merged columns)
    if not col_map.get("donor") or not col_map.get("organizations they picked"):
        df_fb = _read_excel_fallback(excel_path)
        if df_fb is not None and not df_fb.empty:
            df_fb = _normalize_columns(df_fb)
            col_map_fb = _resolve_higgs_columns(df_fb)
            if col_map_fb.get("donor") and col_map_fb.get("organizations they picked"):
                df, col_map = df_fb, col_map_fb
    if not col_map.get("donor") or not col_map.get("organizations they picked"):
        return {"donors_inserted": 0, "ngos_inserted": 0}

    donors_inserted, ngos_inserted = _import_higgs_combined_sheet(df, col_map)
    return {"donors_inserted": donors_inserted, "ngos_inserted": ngos_inserted}


def _read_excel_with_merged_cells_filled(excel_path: Path) -> Optional["pd.DataFrame"]:
    """
    Use openpyxl to fill merged cells (Donor, Legal form) so every row has the value,
    then read into pandas. Handles header row (e.g. 2023) by trying header=0 then header=1.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        # Fallback: pandas only, forward-fill donor and legal form
        return _read_excel_fallback(excel_path)

    wb = load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb.active if wb.active is not None else (wb.worksheets[0] if wb.worksheets else None)
    if ws is None:
        return None

    # Unmerge first (MergedCell is read-only), then fill in a second pass
    merged_ranges = list(ws.merged_cells.ranges)
    ranges_with_values = []
    for merged_range in merged_ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        top_left = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        ranges_with_values.append((min_row, min_col, max_row, max_col, top_left))
    for min_row, min_col, max_row, max_col, top_left in ranges_with_values:
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    wb.close()

    # Try header=1 first (row 2 = real header when row 1 is year), then header=0
    for header_row in (1, 0):
        buffer.seek(0)
        try:
            df = pd.read_excel(buffer, header=header_row, engine="openpyxl")
        except Exception:
            df = pd.read_excel(buffer, header=header_row)
        df = _normalize_columns(df)
        if _has_column(df, "donor") or any("donor" in str(c).lower() for c in df.columns):
            return df
    buffer.seek(0)
    return pd.read_excel(buffer, header=0, engine="openpyxl")


def _read_excel_fallback(excel_path: Path) -> Optional["pd.DataFrame"]:
    """Pandas-only: read and forward-fill Donor / Legal form (merged columns)."""
    # Try header=1 first (row 2 = real header when row 1 is year)
    df = pd.read_excel(excel_path, header=1)
    df = _normalize_columns(df)
    if not _has_column(df, "donor"):
        df = pd.read_excel(excel_path, header=0)
        df = _normalize_columns(df)
    if _has_column(df, "donor"):
        df["donor"] = df["donor"].ffill()
    if _has_column(df, "legal form"):
        df["legal form"] = df["legal form"].ffill()
    return df if _has_column(df, "donor") else None


def _normalize_columns(df: "pd.DataFrame") -> "pd.DataFrame":
    if df.empty:
        return df
    df = df.copy()
    df.columns = [
        str(c).strip().lower() if isinstance(c, str) else c
        for c in df.columns
    ]
    return df


def _has_column(df: "pd.DataFrame", name: str) -> bool:
    return name in df.columns


def _resolve_higgs_columns(df: "pd.DataFrame") -> Dict[str, Optional[str]]:
    """
    Map canonical names to actual column name in df (exact match after normalize, or contains).
    """
    col_list = list(df.columns)
    col_lower = [str(c).strip().lower() for c in col_list]

    def find(must_contain: str) -> Optional[str]:
        for i, c in enumerate(col_lower):
            if must_contain in c or c == must_contain:
                return col_list[i]
        return None

    # Donor strategy = column with "strategy" but not "organization"
    donor_strat = None
    for i, c in enumerate(col_lower):
        if "strategy" in c and "organization" not in c:
            donor_strat = col_list[i]
            break

    return {
        "donor": find("donor"),
        "legal form": find("legal") or find("legal form"),
        "strategy": donor_strat or find("strategy"),
        "organizations they picked": find("organizations they picked") or find("picked") or find("organisation"),
        "organizations strategy": find("organizations strategy") or find("organization strategy"),
        "organization strategy": find("organization strategy") or find("organizations strategy"),
    }


def _safe_str(val: Any) -> Optional[str]:
    """None, NaN, or string 'nan' → None; else stripped string."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    return s


def _import_higgs_combined_sheet(df: "pd.DataFrame", col_map: Dict[str, Optional[str]]) -> tuple[int, int]:
    """
    - Donors: one per unique Donor name (17 for 2023). Legal form = first in block; Strategy = all strategies in block joined.
    - NGOs: one per unique "Organizations they picked", with focus = "Organization strategy" (first occurrence).
    """
    donor_col = col_map.get("donor") or "donor"
    lf_col = col_map.get("legal form") or "legal form"
    strat_col = col_map.get("strategy") or "strategy"
    org_col = col_map.get("organizations they picked") or "organizations they picked"
    org_strat_col = col_map.get("organizations strategy") or col_map.get("organization strategy") or "organizations strategy"

    # Only use columns that exist
    use_cols = [c for c in [donor_col, lf_col, strat_col, org_col, org_strat_col] if c in df.columns]
    df_clean = df.copy()
    for col in use_cols:
        df_clean[col] = df_clean[col].apply(lambda v: _safe_str(v))

    # Drop rows where donor is missing (noise rows)
    if donor_col not in df_clean.columns:
        return 0, 0
    # Keep only rows that have a non-empty donor
    mask = df_clean[donor_col].notna()
    mask = mask & (df_clean[donor_col].astype(str).str.strip() != "")
    mask = mask & (df_clean[donor_col].astype(str).str.lower() != "nan")
    df_clean = df_clean[mask]
    if df_clean.empty:
        return 0, 0

    donors_inserted = 0
    ngos_inserted = 0

    # One donor per unique donor name (e.g. 17 for 2023)
    donor_groups = df_clean.groupby(donor_col, sort=False)
    for donor_name, group in donor_groups:
        name = _safe_str(donor_name)
        if not name:
            continue
        legal_form = None
        if lf_col in group.columns:
            for v in group[lf_col]:
                legal_form = _safe_str(v)
                if legal_form:
                    break
        strategies = []
        if strat_col in group.columns:
            for v in group[strat_col]:
                s = _safe_str(v)
                if s and s not in strategies:
                    strategies.append(s)
        strategy_str = ", ".join(strategies) if strategies else None
        dataset_db.insert_donor(
            name=name,
            legal_form=legal_form,
            strategy=strategy_str,
        )
        donors_inserted += 1

    # One NGO per unique "Organizations they picked"; focus = organization(s) strategy
    seen_ngo = set()
    for _, row in df_clean.iterrows():
        org_name = _safe_str(row.get(org_col))
        if not org_name or org_name in seen_ngo:
            continue
        seen_ngo.add(org_name)
        focus = _safe_str(row.get(org_strat_col)) if org_strat_col in row.index and org_strat_col in df_clean.columns else None
        dataset_db.insert_ngo(
            name=org_name,
            legal_form=None,
            strategy=None,
            focus=focus,
        )
        ngos_inserted += 1

    return donors_inserted, ngos_inserted
