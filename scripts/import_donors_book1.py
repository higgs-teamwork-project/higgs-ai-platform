"""
Import donors from "Book 1" Excel (donors only: name, legal form, strategy).

- Columns A (Donor) and B (Legal form) are MERGED – we fill them so every row has the value.
- Column C (Strategy) has one strategy per cell, no merging.

Usage (from project root, with venv active):
  python scripts/import_donors_book1.py
  python scripts/import_donors_book1.py path/to/book1.xlsx

Does not seed mock data. Only inserts donors from the Excel.
At the end prints the donors so you can check they look right.
"""

import sys
from pathlib import Path

# Project root on path
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

import database
import database.dataset_db as dataset_db


# Default path to Book 1 (donors Excel) in project folder
DEFAULT_BOOK1 = ROOT / "book1.xlsx"

# Column indices (A=1, B=2, C=3)
COL_DONOR = 1
COL_LEGAL_FORM = 2
COL_STRATEGY = 3

# First data row (1-based). Row 1 = year, row 2 = header, row 3 = first data
HEADER_ROW = 2
FIRST_DATA_ROW = 3


def _safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    return s


def import_donors_from_book1(excel_path: Path) -> int:
    """
    Load Book 1 Excel, fill merged cells for A and B, collect strategy from C per donor block.
    Insert one donor per block. Returns number of donors inserted.
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"File not found: {excel_path}")

    database.initialize_all_databases()
    # This script replaces donors with Book 1 contents (no merge with existing)
    conn = dataset_db.get_connection()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM donor_ngo_matches")
        cur.execute("DELETE FROM donors")
        conn.commit()
    finally:
        conn.close()

    wb = load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb.active or wb.worksheets[0]

    # Unmerge and fill merged cells (so every row has donor name and legal form).
    # If top-left is empty (e.g. (PLC) name was in another cell of the merge), use first non-empty in range.
    merged_ranges = list(ws.merged_cells.ranges)
    ranges_with_values = []
    for merged_range in merged_ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        top_left = ws.cell(row=min_row, column=min_col).value
        if _safe_str(top_left) is None:
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    v = ws.cell(row=r, column=c).value
                    if _safe_str(v) is not None:
                        top_left = v
                        break
                if _safe_str(top_left) is not None:
                    break
        ws.unmerge_cells(str(merged_range))
        ranges_with_values.append((min_row, min_col, max_row, max_col, top_left))
    for min_row, min_col, max_row, max_col, top_left in ranges_with_values:
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left

    # Walk rows from FIRST_DATA_ROW; column A = donor, B = legal form, C = strategy (one per cell)
    current_name = None
    current_legal_form = None
    strategies = []
    donors_inserted = 0

    max_row = ws.max_row
    for row in range(FIRST_DATA_ROW, max_row + 1):
        name_val = ws.cell(row=row, column=COL_DONOR).value
        lf_val = ws.cell(row=row, column=COL_LEGAL_FORM).value
        strat_val = ws.cell(row=row, column=COL_STRATEGY).value

        name = _safe_str(name_val)
        legal_form = _safe_str(lf_val)
        strategy = _safe_str(strat_val)

        # New donor block when we have a name in column A
        if name:
            if current_name:
                strategy_str = ", ".join(s for s in strategies if s) if strategies else None
                dataset_db.insert_donor(
                    name=current_name,
                    legal_form=current_legal_form,
                    strategy=strategy_str,
                )
                donors_inserted += 1
            current_name = name
            current_legal_form = legal_form
            strategies = [strategy] if strategy else []
        else:
            # Same donor block: collect strategy from this row (column C)
            if strategy and current_name:
                strategies.append(strategy)

    # Last donor
    if current_name:
        strategy_str = ", ".join(s for s in strategies if s) if strategies else None
        dataset_db.insert_donor(
            name=current_name,
            legal_form=current_legal_form,
            strategy=strategy_str,
        )
        donors_inserted += 1

    wb.close()
    return donors_inserted


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_BOOK1
    if not path.is_absolute():
        path = ROOT / path

    print(f"Using: {path}")
    if not path.exists():
        print(f"ERROR: File not found. Put your donors Excel at {path} or run:")
        print(f"  python scripts/import_donors_book1.py <path_to_book1.xlsx>")
        sys.exit(1)

    n = import_donors_from_book1(path)
    print(f"\nDonors inserted: {n}")

    donors = dataset_db.list_donors()
    print("\n--- Donors in DB (check that they look right) ---")
    for row in donors:
        name = row["name"]
        lf = row["legal_form"] or ""
        strat = (row["strategy"] or "")[:60]
        if len(row["strategy"] or "") > 60:
            strat += "..."
        print(f"  [{row['id']}] {name}  |  {lf}  |  {strat}")
    print("---")


if __name__ == "__main__":
    main()
