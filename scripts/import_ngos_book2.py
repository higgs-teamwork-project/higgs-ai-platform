"""
Import NGOs from "Book 2" Excel (NGOs only: name, strategy/focus).

- Column A: NGO name (one per row, no merged cells).
- Columns B and C: strategy/focus (each cell can have comma-separated terms; we combine B and C per row).

Usage (from project root, with venv active):
  python scripts/import_ngos_book2.py
  python scripts/import_ngos_book2.py path/to/book2.xlsx

Replaces existing NGOs and matches. At the end prints the NGOs so you can check they look right.
"""

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

import database
import database.dataset_db as dataset_db


DEFAULT_BOOK2 = ROOT / "book2.xlsx"

COL_NAME = 1
COL_STRATEGY_B = 2
COL_STRATEGY_C = 3

# Row 1 = header, row 2 = first data
FIRST_DATA_ROW = 2


def _safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    return s


def _combine_strategy(b_val, c_val):
    """Combine B and C into one string; keep comma-separated terms as-is."""
    parts = []
    for v in (b_val, c_val):
        s = _safe_str(v)
        if s:
            parts.append(s)
    if not parts:
        return None
    return ", ".join(parts)


def import_ngos_from_book2(excel_path: Path) -> int:
    """
    Load Book 2 Excel: one row per NGO, A=name, B+C=strategy/focus.
    Replaces existing NGOs (and matches). Returns number inserted.
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"File not found: {excel_path}")

    database.initialize_all_databases()

    conn = dataset_db.get_connection()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM donor_ngo_matches")
        cur.execute("DELETE FROM ngos")
        conn.commit()
    finally:
        conn.close()

    wb = load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb.active or wb.worksheets[0]

    # Only add unique NGOs by name (first occurrence wins; merge strategy if we see same name again)
    seen_names = set()
    inserted = 0
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        name_val = ws.cell(row=row, column=COL_NAME).value
        b_val = ws.cell(row=row, column=COL_STRATEGY_B).value
        c_val = ws.cell(row=row, column=COL_STRATEGY_C).value

        name = _safe_str(name_val)
        if not name:
            continue

        name_key = name.strip().lower()
        if name_key in seen_names:
            continue
        seen_names.add(name_key)

        strategy = _combine_strategy(b_val, c_val)
        dataset_db.insert_ngo(
            name=name,
            strategy=strategy,
            focus=strategy,
        )
        inserted += 1

    wb.close()
    return inserted


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_BOOK2
    if not path.is_absolute():
        path = ROOT / path

    print(f"Using: {path}")
    if not path.exists():
        print(f"ERROR: File not found. Put your NGOs Excel at {path} or run:")
        print(f"  python scripts/import_ngos_book2.py <path_to_book2.xlsx>")
        sys.exit(1)

    n = import_ngos_from_book2(path)
    print(f"\nNGOs inserted: {n}")

    ngos = dataset_db.list_ngos()
    print("\n--- NGOs in DB (check that they look right) ---")
    for row in ngos:
        name = row["name"]
        strat = (row["strategy"] or row["focus"] or "")[:60]
        if len(row["strategy"] or row["focus"] or "") > 60:
            strat += "..."
        print(f"  [{row['id']}] {name}  |  {strat}")
    print("---")


if __name__ == "__main__":
    main()
